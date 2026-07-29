"""
MÓDULO — CONTROLE DE MATERIAIS
• NLAG  → saldo controlado pelo sistema (entradas/saídas/inventário)
• HIBE/ERSA → saldo importado do SAP (somente consulta de disponibilidade)
Inclui etiquetas com código de barras, coletor, alertas de estoque mínimo
e relatório semanal para o comprador.
"""
import io
import csv
import base64

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, session, Response, jsonify, abort)
import psycopg2

import db
from auth import exige, pode

bp = Blueprint("mat", __name__, url_prefix="/materiais")

try:
    import barcode as python_barcode
    from barcode.writer import ImageWriter
    from PIL import Image, ImageChops
    BARCODE_OK = True
except Exception:
    BARCODE_OK = False


# ══════════════════════════════════════════════════════════════════
#  SALDO / DASHBOARD
# ══════════════════════════════════════════════════════════════════
@bp.route("/")
@exige("material_ver")
def saldo():
    tipo = request.args.get("tipo", "NLAG")
    busca = request.args.get("q", "").strip()
    filtro = request.args.get("filtro", "")

    where = ["m.ativo=TRUE"]
    params = []
    if tipo != "todos":
        where.append("m.tipo=%s")
        params.append(tipo)
    if busca:
        where.append("(m.codigo ILIKE %s OR m.descricao ILIKE %s)")
        params += [f"%{busca}%", f"%{busca}%"]

    itens = db.query(f"""
        SELECT m.*,
          COALESCE(SUM(CASE WHEN mv.tipo IN ('ENTRADA','AJUSTE') THEN mv.quantidade
                            ELSE -mv.quantidade END),0) AS saldo_nlag
        FROM materiais m
        LEFT JOIN movimentacoes mv ON mv.codigo=m.codigo
        WHERE {' AND '.join(where)}
        GROUP BY m.id
        ORDER BY m.codigo""", params)

    # saldo efetivo conforme o tipo
    linhas = []
    for m in itens or []:
        m = dict(m)
        m["saldo"] = float(m["saldo_nlag"]) if m["tipo"] == "NLAG" else float(m["saldo_sap"] or 0)
        m["abaixo_min"] = m["estoque_min"] and m["saldo"] < float(m["estoque_min"])
        if filtro == "criticos" and not m["abaixo_min"]:
            continue
        if filtro == "zerados" and m["saldo"] > 0:
            continue
        linhas.append(m)

    resumo = {
        "total": len(linhas),
        "abaixo": len([m for m in linhas if m["abaixo_min"]]),
        "zerados": len([m for m in linhas if m["saldo"] <= 0]),
    }
    return render_template("mat/saldo.html", itens=linhas, tipo=tipo, busca=busca,
                           filtro=filtro, resumo=resumo)


# ══════════════════════════════════════════════════════════════════
#  CADASTRO
# ══════════════════════════════════════════════════════════════════
@bp.route("/cadastro", methods=["GET", "POST"])
@exige("material_cad")
def cadastro():
    if request.method == "POST":
        acao = request.form.get("acao", "novo")
        codigo = request.form.get("codigo", "").strip().upper()
        if not codigo:
            flash("Informe o código do material.", "warning")
            return redirect(url_for("mat.cadastro"))

        imagem = None
        f = request.files.get("imagem")
        if f and f.filename:
            imagem = _processar_imagem(f)

        if acao == "editar":
            sql = """UPDATE materiais SET descricao=%s, unidade=%s, tipo=%s, aplicacao=%s,
                     critico=%s, estoque_min=%s, estoque_max=%s, valor_unit=%s,
                     localizacao=%s, ativo=%s, atualizado_em=NOW()"""
            params = [request.form["descricao"].strip(), request.form.get("unidade", "UNI"),
                      request.form.get("tipo", "NLAG"), request.form.get("aplicacao", "").strip(),
                      request.form.get("critico") == "1",
                      request.form.get("estoque_min") or 0, request.form.get("estoque_max") or 0,
                      request.form.get("valor_unit") or 0,
                      request.form.get("localizacao", "").strip(),
                      request.form.get("ativo", "1") == "1"]
            if imagem:
                sql += ", imagem=%s"
                params.append(psycopg2.Binary(imagem))
            sql += " WHERE codigo=%s"
            params.append(codigo)
            db.executar(sql, params)
            flash("Material atualizado.", "success")
        else:
            db.executar("""INSERT INTO materiais
                  (codigo, descricao, unidade, tipo, aplicacao, critico, estoque_min,
                   estoque_max, valor_unit, localizacao, imagem)
                  VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                  ON CONFLICT (codigo) DO UPDATE SET descricao=EXCLUDED.descricao,
                    unidade=EXCLUDED.unidade, atualizado_em=NOW()""",
                (codigo, request.form["descricao"].strip(), request.form.get("unidade", "UNI"),
                 request.form.get("tipo", "NLAG"), request.form.get("aplicacao", "").strip(),
                 request.form.get("critico") == "1",
                 request.form.get("estoque_min") or 0, request.form.get("estoque_max") or 0,
                 request.form.get("valor_unit") or 0,
                 request.form.get("localizacao", "").strip(),
                 psycopg2.Binary(imagem) if imagem else None))
            flash("Material cadastrado.", "success")
        return redirect(url_for("mat.cadastro"))

    busca = request.args.get("q", "").strip()
    editar = request.args.get("editar", "").strip().upper()
    where, params = "ativo=TRUE OR ativo=FALSE", []
    if busca:
        where = "(codigo ILIKE %s OR descricao ILIKE %s)"
        params = [f"%{busca}%", f"%{busca}%"]
    itens = db.query(f"SELECT * FROM materiais WHERE {where} ORDER BY codigo LIMIT 300", params)
    material = db.um("SELECT * FROM materiais WHERE codigo=%s", (editar,)) if editar else None
    return render_template("mat/cadastro.html", itens=itens, busca=busca, material=material)


def _processar_imagem(f):
    try:
        from PIL import Image
        img = Image.open(f.stream).convert("RGB")
        img.thumbnail((800, 800))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=82, optimize=True)
        return buf.getvalue()
    except Exception:
        return None


@bp.route("/imagem/<codigo>")
def imagem(codigo):
    r = db.um("SELECT imagem FROM materiais WHERE codigo=%s", (codigo.strip().upper(),))
    if r and r["imagem"]:
        return Response(bytes(r["imagem"]), mimetype="image/jpeg")
    svg = ('<svg xmlns="http://www.w3.org/2000/svg" width="80" height="80">'
           '<rect width="80" height="80" fill="#EEF1F6"/>'
           '<text x="40" y="44" font-size="9" fill="#8B94A3" text-anchor="middle">sem foto</text>'
           '</svg>')
    return Response(svg, mimetype="image/svg+xml")


# ══════════════════════════════════════════════════════════════════
#  MOVIMENTAÇÕES
# ══════════════════════════════════════════════════════════════════
@bp.route("/entrada", methods=["GET", "POST"])
@exige("material_mov")
def entrada():
    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        qtd = float(request.form.get("quantidade") or 0)
        obs = request.form.get("observacao", "").strip()
        m = db.um("SELECT * FROM materiais WHERE codigo=%s", (codigo,))
        if not m:
            flash(f"Material {codigo} não cadastrado.", "danger")
        elif qtd <= 0:
            flash("Quantidade inválida.", "warning")
        else:
            db.executar("""INSERT INTO movimentacoes (codigo, tipo, quantidade, usuario, observacao)
                           VALUES (%s,'ENTRADA',%s,%s,%s)""",
                        (codigo, qtd, session["nome"], obs))
            flash(f"Entrada de {qtd:g} {m['unidade']} — {m['descricao']}.", "success")
        return redirect(url_for("mat.entrada"))

    ultimas = db.query("""SELECT mv.*, m.descricao, m.unidade FROM movimentacoes mv
                          LEFT JOIN materiais m ON m.codigo=mv.codigo
                          WHERE mv.tipo='ENTRADA' ORDER BY mv.data_hora DESC LIMIT 15""")
    return render_template("mat/entrada.html", ultimas=ultimas)


@bp.route("/saida", methods=["GET", "POST"])
@exige("material_mov")
def saida():
    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        qtd = float(request.form.get("quantidade") or 0)
        obs = request.form.get("observacao", "").strip()
        os_id = request.form.get("os_id") or None
        m = db.um("SELECT * FROM materiais WHERE codigo=%s", (codigo,))
        if not m:
            flash(f"Material {codigo} não cadastrado.", "danger")
        elif qtd <= 0:
            flash("Quantidade inválida.", "warning")
        else:
            saldo = db.saldo_material(codigo)
            if saldo < qtd:
                flash(f"Saldo insuficiente. Disponível: {saldo:g} {m['unidade']}.", "danger")
            else:
                db.executar("""INSERT INTO movimentacoes
                               (codigo, tipo, quantidade, usuario, observacao, os_id)
                               VALUES (%s,'SAIDA',%s,%s,%s,%s)""",
                            (codigo, qtd, session["nome"], obs, os_id))
                novo = saldo - qtd
                flash(f"Saída registrada. Novo saldo: {novo:g} {m['unidade']}.", "success")
                _alerta_minimo(m, novo)
        return redirect(url_for("mat.saida"))

    ultimas = db.query("""SELECT mv.*, m.descricao, m.unidade FROM movimentacoes mv
                          LEFT JOIN materiais m ON m.codigo=mv.codigo
                          WHERE mv.tipo='SAIDA' ORDER BY mv.data_hora DESC LIMIT 15""")
    abertas = db.query("""SELECT id, numero FROM ordens_servico
                          WHERE status IN ('em_andamento','aguardando_peca','atribuida','aberta')
                          ORDER BY numero DESC LIMIT 50""")
    return render_template("mat/saida.html", ultimas=ultimas, abertas=abertas)


def _alerta_minimo(material, saldo_novo):
    if material["estoque_min"] and saldo_novo < float(material["estoque_min"]):
        sugestao = float(material["estoque_max"] or 0) - saldo_novo
        db.notificar_perfis(
            ("analista", "lider", "supervisao"),
            f"Estoque mínimo atingido — {material['codigo']}",
            f"{material['descricao']} — saldo {saldo_novo:g} {material['unidade']} "
            f"(mín. {float(material['estoque_min']):g}). "
            f"Sugestão de compra: {max(sugestao, 0):g}.",
            url_for("mat.alertas"))


@bp.route("/ajuste", methods=["POST"])
@exige("material_cad")
def ajuste():
    """Inventário: ajusta o saldo para o valor contado."""
    codigo = request.form.get("codigo", "").strip().upper()
    contado = float(request.form.get("contado") or 0)
    m = db.um("SELECT * FROM materiais WHERE codigo=%s", (codigo,))
    if not m:
        flash("Material não encontrado.", "danger")
        return redirect(url_for("mat.saldo"))
    atual = db.saldo_material(codigo)
    diff = contado - atual
    if diff != 0:
        db.executar("""INSERT INTO movimentacoes (codigo, tipo, quantidade, usuario, observacao)
                       VALUES (%s,'AJUSTE',%s,%s,%s)""",
                    (codigo, diff, session["nome"],
                     f"Inventário: de {atual:g} para {contado:g}"))
    flash(f"Saldo de {codigo} ajustado para {contado:g}.", "success")
    return redirect(url_for("mat.saldo"))


@bp.route("/historico")
@exige("material_ver")
def historico():
    codigo = request.args.get("codigo", "").strip().upper()
    tipo = request.args.get("tipo", "")
    where, params = ["1=1"], []
    if codigo:
        where.append("mv.codigo=%s")
        params.append(codigo)
    if tipo:
        where.append("mv.tipo=%s")
        params.append(tipo)
    itens = db.query(f"""SELECT mv.*, m.descricao, m.unidade, o.numero AS os_numero
                         FROM movimentacoes mv
                         LEFT JOIN materiais m ON m.codigo=mv.codigo
                         LEFT JOIN ordens_servico o ON o.id=mv.os_id
                         WHERE {' AND '.join(where)}
                         ORDER BY mv.data_hora DESC LIMIT 500""", params)
    return render_template("mat/historico.html", itens=itens, codigo=codigo, tipo=tipo)


# ══════════════════════════════════════════════════════════════════
#  COLETOR (leitor de código de barras)
# ══════════════════════════════════════════════════════════════════
@bp.route("/coletor", methods=["GET", "POST"])
@exige("material_mov")
def coletor():
    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        qtd = float(request.form.get("quantidade") or 1)
        operacao = request.form.get("operacao", "SAIDA")
        m = db.um("SELECT * FROM materiais WHERE codigo=%s", (codigo,))
        if not m:
            return jsonify(ok=False, msg=f"Material {codigo} não cadastrado.")
        if operacao == "SAIDA":
            saldo = db.saldo_material(codigo)
            if saldo < qtd:
                return jsonify(ok=False, msg=f"Saldo insuficiente ({saldo:g}).")
        db.executar("""INSERT INTO movimentacoes (codigo, tipo, quantidade, usuario, observacao)
                       VALUES (%s,%s,%s,%s,'Coletor')""",
                    (codigo, operacao, qtd, session["nome"]))
        novo = db.saldo_material(codigo)
        if operacao == "SAIDA":
            _alerta_minimo(m, novo)
        return jsonify(ok=True, msg=f"{m['descricao']} — saldo {novo:g} {m['unidade']}",
                       descricao=m["descricao"], saldo=novo)
    return render_template("mat/coletor.html")


# ══════════════════════════════════════════════════════════════════
#  ETIQUETAS COM CÓDIGO DE BARRAS
# ══════════════════════════════════════════════════════════════════
def gerar_barcode_b64(codigo):
    if not BARCODE_OK:
        return None
    try:
        opts = {"module_width": 0.3, "module_height": 10.0, "font_size": 0,
                "text_distance": 0, "quiet_zone": 2.0, "dpi": 300, "write_text": False}
        c128 = python_barcode.get("code128", str(codigo), writer=ImageWriter())
        buf = io.BytesIO()
        c128.write(buf, options=opts)
        buf.seek(0)
        img = Image.open(buf).convert("RGB")
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bbox = ImageChops.difference(img, bg).getbbox()
        if bbox:
            img = img.crop(bbox)
        razao = img.width / img.height
        img = img.resize((int(85 * razao), 85), Image.LANCZOS)
        out = io.BytesIO()
        img.save(out, format="PNG", optimize=True)
        return base64.b64encode(out.getvalue()).decode()
    except Exception as e:
        print(f"[barcode] {e}", flush=True)
        return None


@bp.route("/etiquetas")
@exige("material_ver")
def etiquetas():
    busca = request.args.get("q", "").strip()
    itens = []
    if busca:
        itens = db.query("""SELECT * FROM materiais
                            WHERE (codigo ILIKE %s OR descricao ILIKE %s) AND ativo=TRUE
                            ORDER BY codigo LIMIT 100""", (f"%{busca}%", f"%{busca}%"))
    return render_template("mat/etiquetas.html", itens=itens, busca=busca)


@bp.route("/etiqueta/imprimir")
@exige("material_ver")
def etiqueta_imprimir():
    codigos = request.args.getlist("codigo")
    copias = int(request.args.get("copias", 1))
    etiq = []
    for c in codigos:
        m = db.um("SELECT * FROM materiais WHERE codigo=%s", (c.strip().upper(),))
        if m:
            b64 = gerar_barcode_b64(m["codigo"])
            for _ in range(copias):
                etiq.append({"codigo": m["codigo"], "descricao": m["descricao"],
                             "unidade": m["unidade"], "barcode": b64,
                             "localizacao": m["localizacao"]})
    return render_template("mat/etiqueta_print.html", etiquetas=etiq)


# ══════════════════════════════════════════════════════════════════
#  IMPORTAÇÃO / EXPORTAÇÃO
# ══════════════════════════════════════════════════════════════════
@bp.route("/importar", methods=["GET", "POST"])
@exige("material_cad")
def importar():
    if request.method == "POST":
        destino = request.form.get("destino", "cadastro")
        f = request.files.get("arquivo")
        if not f or not f.filename:
            flash("Selecione um arquivo.", "warning")
            return redirect(url_for("mat.importar"))

        nome = f.filename.lower()
        try:
            if nome.endswith((".xlsx", ".xlsm", ".xls")):
                linhas = _ler_excel(f)
            else:
                linhas = _ler_csv(f)
        except Exception as e:
            flash(f"Não foi possível ler o arquivo: {e}", "danger")
            return redirect(url_for("mat.importar"))

        if destino == "sap":
            n = _importar_saldo_sap(linhas)
            flash(f"{n} saldo(s) HIBE/ERSA atualizado(s) a partir do SAP.", "success")
        elif destino == "nlag":
            n = _importar_saldo_nlag(linhas)
            flash(f"{n} saldo(s) NLAG ajustado(s).", "success")
        else:
            n = _importar_cadastro(linhas)
            flash(f"{n} material(is) cadastrado(s)/atualizado(s).", "success")
        return redirect(url_for("mat.saldo"))

    return render_template("mat/importar.html")


def _ler_csv(f):
    conteudo = f.read().decode("utf-8-sig", errors="ignore")
    dialeto = ";" if conteudo.count(";") > conteudo.count(",") else ","
    return list(csv.DictReader(io.StringIO(conteudo), delimiter=dialeto))


def _ler_excel(f):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas, cabecalho = [], None
    for row in ws.iter_rows(values_only=True):
        vals = [("" if c is None else str(c).strip()) for c in row]
        if not any(vals):
            continue
        if cabecalho is None:
            cabecalho = vals
            continue
        linhas.append(dict(zip(cabecalho, vals)))
    wb.close()
    return linhas


def _campo(linha, *nomes):
    """Busca o valor por vários possíveis nomes de coluna (case-insensitive)."""
    norm = {str(k).strip().lower(): v for k, v in linha.items() if k}
    for n in nomes:
        v = norm.get(n.lower())
        if v not in (None, ""):
            return str(v).strip()
    return ""


def _num(v):
    if not v:
        return 0.0
    v = str(v).upper()
    for lixo in ("PEÇ", "UNI", "UN", "PC", "KG", "L", "M", "\xa0"):
        v = v.replace(lixo, "")
    v = v.strip().replace(".", "").replace(",", ".") if v.count(",") == 1 and v.count(".") > 1 \
        else v.strip().replace(",", ".")
    try:
        return float(v)
    except Exception:
        return 0.0


def _importar_cadastro(linhas):
    n = 0
    for l in linhas:
        codigo = _campo(l, "codigo", "código", "material", "cod")
        if not codigo:
            continue
        desc = _campo(l, "descricao", "descrição", "descrição de material", "descricao de material")
        db.executar("""INSERT INTO materiais (codigo, descricao, unidade, tipo, estoque_min, estoque_max)
                       VALUES (%s,%s,%s,%s,%s,%s)
                       ON CONFLICT (codigo) DO UPDATE
                       SET descricao=COALESCE(NULLIF(EXCLUDED.descricao,''), materiais.descricao),
                           atualizado_em=NOW()""",
                    (codigo.upper(), desc or codigo,
                     _campo(l, "unidade", "umb", "um") or "UNI",
                     _campo(l, "tipo") or "NLAG",
                     _num(_campo(l, "min", "mín", "estoque_min")),
                     _num(_campo(l, "max", "máx", "estoque_max"))))
        n += 1
    return n


def _importar_saldo_sap(linhas):
    """Planilha do SAP: Material / Descrição / Estoque de utilização livre."""
    n = 0
    for l in linhas:
        codigo = _campo(l, "material", "codigo", "código")
        if not codigo:
            continue
        saldo = _num(_campo(l, "estoque de utilização livre", "estoque de utilizacao livre",
                            "saldo", "estoque livre", "estoque"))
        desc = _campo(l, "descrição de material", "descricao de material", "descricao", "descrição")
        db.executar("""INSERT INTO materiais (codigo, descricao, tipo, saldo_sap, atualizado_em)
                       VALUES (%s,%s,'HIBE',%s,NOW())
                       ON CONFLICT (codigo) DO UPDATE
                       SET saldo_sap=EXCLUDED.saldo_sap, atualizado_em=NOW(),
                           descricao=COALESCE(NULLIF(EXCLUDED.descricao,''), materiais.descricao)""",
                    (codigo.upper(), desc or codigo, saldo))
        n += 1
    return n


def _importar_saldo_nlag(linhas):
    """Planilha SALDO NLAG: Codigo / Descricao / Unidade / Saldo → gera ajuste."""
    n = 0
    for l in linhas:
        codigo = _campo(l, "codigo", "código", "material")
        if not codigo:
            continue
        codigo = codigo.upper()
        saldo = _num(_campo(l, "saldo", "quantidade", "qtd"))
        desc = _campo(l, "descricao", "descrição")
        db.executar("""INSERT INTO materiais (codigo, descricao, unidade, tipo)
                       VALUES (%s,%s,%s,'NLAG')
                       ON CONFLICT (codigo) DO UPDATE
                       SET descricao=COALESCE(NULLIF(EXCLUDED.descricao,''), materiais.descricao)""",
                    (codigo, desc or codigo, _campo(l, "unidade", "umb") or "UNI"))
        atual = db.saldo_material(codigo)
        diff = saldo - atual
        if diff != 0:
            db.executar("""INSERT INTO movimentacoes (codigo, tipo, quantidade, usuario, observacao)
                           VALUES (%s,'AJUSTE',%s,%s,'Importação de saldo NLAG')""",
                        (codigo, diff, session.get("nome", "sistema")))
        n += 1
    return n


@bp.route("/exportar")
@exige("material_ver")
def exportar():
    itens = db.query("""SELECT m.codigo, m.descricao, m.unidade, m.tipo, m.estoque_min,
                               m.estoque_max, m.saldo_sap, m.localizacao,
              COALESCE(SUM(CASE WHEN mv.tipo IN ('ENTRADA','AJUSTE') THEN mv.quantidade
                                ELSE -mv.quantidade END),0) AS saldo_nlag
              FROM materiais m LEFT JOIN movimentacoes mv ON mv.codigo=m.codigo
              WHERE m.ativo=TRUE GROUP BY m.id ORDER BY m.codigo""")
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Codigo", "Descricao", "Unidade", "Tipo", "Saldo", "Minimo", "Maximo", "Local"])
    for m in itens or []:
        saldo = m["saldo_nlag"] if m["tipo"] == "NLAG" else m["saldo_sap"]
        w.writerow([m["codigo"], m["descricao"], m["unidade"], m["tipo"],
                    f"{float(saldo or 0):g}", f"{float(m['estoque_min'] or 0):g}",
                    f"{float(m['estoque_max'] or 0):g}", m["localizacao"] or ""])
    return Response(buf.getvalue().encode("utf-8-sig"), mimetype="text/csv",
                    headers={"Content-Disposition":
                             "attachment; filename=saldo_materiais.csv"})


# ══════════════════════════════════════════════════════════════════
#  ALERTAS / RELATÓRIO DO COMPRADOR
# ══════════════════════════════════════════════════════════════════
@bp.route("/alertas")
@exige("material_ver")
def alertas():
    itens = db.query("""
        SELECT m.*,
          COALESCE(SUM(CASE WHEN mv.tipo IN ('ENTRADA','AJUSTE') THEN mv.quantidade
                            ELSE -mv.quantidade END),0) AS saldo_nlag
        FROM materiais m LEFT JOIN movimentacoes mv ON mv.codigo=m.codigo
        WHERE m.ativo=TRUE AND m.estoque_min > 0
        GROUP BY m.id ORDER BY m.critico DESC, m.codigo""")

    linhas = []
    for m in itens or []:
        m = dict(m)
        saldo = float(m["saldo_nlag"]) if m["tipo"] == "NLAG" else float(m["saldo_sap"] or 0)
        minimo = float(m["estoque_min"] or 0)
        if saldo >= minimo:
            continue
        maximo = float(m["estoque_max"] or 0)
        m["saldo"] = saldo
        m["sugestao"] = round(max(maximo - saldo, minimo - saldo), 2)
        m["consumo_30d"] = float(db.scalar("""SELECT COALESCE(SUM(quantidade),0) AS q
                                              FROM movimentacoes WHERE codigo=%s AND tipo='SAIDA'
                                                AND data_hora > NOW()-INTERVAL '30 days'""",
                                           (m["codigo"],), default=0))
        linhas.append(m)
    linhas.sort(key=lambda x: (not x["critico"], x["saldo"] - float(x["estoque_min"])))
    return render_template("mat/alertas.html", itens=linhas)


@bp.route("/alertas/enviar", methods=["POST"])
@exige("material_cad")
def enviar_alertas():
    """Dispara o alerta semanal para comprador, analista, líder e supervisão."""
    n = db.scalar("""SELECT COUNT(*) AS n FROM (
        SELECT m.codigo FROM materiais m LEFT JOIN movimentacoes mv ON mv.codigo=m.codigo
        WHERE m.ativo=TRUE AND m.estoque_min>0 GROUP BY m.codigo, m.estoque_min, m.tipo, m.saldo_sap
        HAVING (CASE WHEN m.tipo='NLAG' THEN
                  COALESCE(SUM(CASE WHEN mv.tipo IN ('ENTRADA','AJUSTE') THEN mv.quantidade
                                    ELSE -mv.quantidade END),0)
                ELSE m.saldo_sap END) < m.estoque_min) t""")
    db.notificar_perfis(("analista", "lider", "supervisao", "admin"),
                        "Relatório semanal de estoque mínimo",
                        f"{n} item(ns) abaixo do estoque mínimo, com sugestão de compra.",
                        url_for("mat.alertas"))
    flash(f"Alerta enviado — {n} item(ns) abaixo do mínimo.", "success")
    return redirect(url_for("mat.alertas"))


# ══════════════════════════════════════════════════════════════════
#  API interna
# ══════════════════════════════════════════════════════════════════
@bp.route("/api/<codigo>")
def api_material(codigo):
    m = db.um("SELECT * FROM materiais WHERE codigo=%s", (codigo.strip().upper(),))
    if not m:
        return jsonify(ok=False)
    saldo = db.saldo_material(m["codigo"]) if m["tipo"] == "NLAG" else float(m["saldo_sap"] or 0)
    return jsonify(ok=True, codigo=m["codigo"], descricao=m["descricao"],
                   unidade=m["unidade"], tipo=m["tipo"], saldo=saldo,
                   minimo=float(m["estoque_min"] or 0))
