"""
CENTRAL DE RELATÓRIOS E BACKUP
Exportação em Excel (.xlsx) de todos os módulos e backup completo da base.

Observação técnica: as planilhas são geradas no servidor, onde não há Excel
nem LibreOffice para recalcular. Por isso os relatórios trazem valores já
apurados, nunca fórmulas — uma fórmula escrita aqui chegaria vazia ao usuário.
"""
import io
import json
from datetime import date, datetime, timedelta
from decimal import Decimal

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, session, Response, abort)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
from auth import exige, pode

bp = Blueprint("rel", __name__, url_prefix="/relatorios")

# ── Identidade visual da Décio nas planilhas ──
AZUL = "10477D"
VERDE = "28A353"
CINZA = "F2F5FA"
FONTE = "Arial"

F_TITULO = Font(name=FONTE, size=14, bold=True, color="FFFFFF")
F_SUB = Font(name=FONTE, size=9, color="FFFFFF")
F_CAB = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
F_DADO = Font(name=FONTE, size=10)
F_TOTAL = Font(name=FONTE, size=10, bold=True)
P_TITULO = PatternFill("solid", fgColor=AZUL)
P_CAB = PatternFill("solid", fgColor=VERDE)
P_ZEBRA = PatternFill("solid", fgColor=CINZA)
BORDA = Border(bottom=Side(style="thin", color="D0D7E2"))


def _valor(v):
    """Normaliza o valor para algo que o openpyxl aceite."""
    if isinstance(v, datetime):
        # O Excel não aceita datetime com fuso horário
        try:
            return v.astimezone(db.TZ_BR).replace(tzinfo=None)
        except (ValueError, TypeError):
            return v.replace(tzinfo=None)
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (bytes, memoryview)):
        return "[binário]"
    if isinstance(v, bool):
        return "Sim" if v else "Não"
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False, default=str)
    return v


def escrever_aba(ws, titulo, colunas, linhas, subtitulo="", larguras=None,
                 formatos=None, totalizar=None):
    """
    Monta uma aba padronizada:
      linha 1 → título   ·   linha 2 → subtítulo   ·   linha 4 → cabeçalho
      linha 5+ → dados   ·   última linha → totais (opcional)
    formatos: {índice_da_coluna: 'formato_numérico'}
    totalizar: lista de índices de coluna que recebem soma no rodapé
    """
    n = max(len(colunas), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font, c.fill = F_TITULO, P_TITULO
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=subtitulo or
                f"Décio Metalúrgica · gerado em {db.agora():%d/%m/%Y %H:%M}")
    c.font, c.fill = F_SUB, P_TITULO
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    for i, nome in enumerate(colunas, 1):
        c = ws.cell(row=4, column=i, value=nome)
        c.font, c.fill = F_CAB, P_CAB
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 24

    formatos = formatos or {}
    for r, linha in enumerate(linhas, 5):
        for i, v in enumerate(linha, 1):
            c = ws.cell(row=r, column=i, value=_valor(v))
            c.font = F_DADO
            c.border = BORDA
            if r % 2 == 1:
                c.fill = P_ZEBRA
            if i in formatos:
                c.number_format = formatos[i]
            elif isinstance(v, datetime):
                c.number_format = "DD/MM/YYYY HH:MM"
            elif isinstance(v, date):
                c.number_format = "DD/MM/YYYY"

    if totalizar and linhas:
        r = len(linhas) + 5
        ws.cell(row=r, column=1, value="TOTAL").font = F_TOTAL
        for i in totalizar:
            soma = 0.0
            for linha in linhas:
                try:
                    soma += float(linha[i - 1] or 0)
                except (TypeError, ValueError, IndexError):
                    pass
            c = ws.cell(row=r, column=i, value=round(soma, 2))
            c.font = F_TOTAL
            c.number_format = formatos.get(i, "#,##0.00")

    # Largura das colunas
    larguras = larguras or []
    for i in range(1, n + 1):
        if i <= len(larguras) and larguras[i - 1]:
            largura = larguras[i - 1]
        else:
            maior = len(str(colunas[i - 1])) if i <= len(colunas) else 10
            for linha in linhas[:200]:
                if i <= len(linha) and linha[i - 1] is not None:
                    maior = max(maior, min(len(str(linha[i - 1])), 55))
            largura = maior + 3
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A5"
    if linhas:
        ws.auto_filter.ref = f"A4:{get_column_letter(n)}{len(linhas) + 4}"
    return ws


def enviar(wb, nome):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    arquivo = f"{nome}_{db.agora():%Y%m%d_%H%M}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={arquivo}"})


def periodo():
    """Lê ini/fim da querystring; padrão: últimos 90 dias."""
    fim = request.args.get("fim") or db.hoje().isoformat()
    ini = request.args.get("ini") or (db.hoje() - timedelta(days=90)).isoformat()
    return ini, fim


def rotulo_periodo(ini, fim):
    return (f"Décio Metalúrgica · período de "
            f"{date.fromisoformat(ini):%d/%m/%Y} a {date.fromisoformat(fim):%d/%m/%Y} · "
            f"gerado em {db.agora():%d/%m/%Y %H:%M}")


MOEDA = 'R$ #,##0.00'
NUM = '#,##0.###'


# ══════════════════════════════════════════════════════════════════
#  PÁGINA INICIAL
# ══════════════════════════════════════════════════════════════════
@bp.route("/")
@exige("relatorios")
def index():
    ini, fim = periodo()
    contagens = {
        "os": db.scalar("SELECT COUNT(*) AS n FROM ordens_servico"),
        "om": db.scalar("SELECT COUNT(*) AS n FROM ordens_manutencao"),
        "mov": db.scalar("SELECT COUNT(*) AS n FROM movimentacoes"),
        "sm": db.scalar("SELECT COUNT(*) AS n FROM solicitacoes_material"),
        "mat": db.scalar("SELECT COUNT(*) AS n FROM materiais"),
        "eq": db.scalar("SELECT COUNT(*) AS n FROM equipamentos"),
    }
    return render_template("rel/index.html", ini=ini, fim=fim, contagens=contagens)


# ══════════════════════════════════════════════════════════════════
#  CORRETIVAS
# ══════════════════════════════════════════════════════════════════
@bp.route("/os")
@exige("relatorios")
def rel_os():
    ini, fim = periodo()
    dados = db.query("""
        SELECT o.numero, o.data_abertura, est.nome AS estab, o.tipo,
               ct.nome AS setor, e.codigo AS eq_codigo,
               COALESCE(e.nome, o.equipamento_outro) AS equipamento,
               o.criticidade, o.tipo_manutencao, o.origem, o.descricao_problema,
               s.nome AS solicitante, r.nome AS responsavel, o.status,
               o.maquina_parada, o.data_inicio, o.data_conclusao, o.data_aprovacao,
               d.nome AS defeito, c.nome AS causa, o.acao_realizada,
               o.tempo_trabalho_seg, o.custo_pecas, o.custo_hh, o.comentario_reprova
        FROM ordens_servico o
        LEFT JOIN equipamentos e ON e.id=o.equipamento_id
        LEFT JOIN centros_trabalho ct ON ct.id=o.centro_trabalho_id
        LEFT JOIN estabelecimentos est ON est.id=o.estabelecimento_id
        LEFT JOIN usuarios s ON s.id=o.solicitante_id
        LEFT JOIN usuarios r ON r.id=o.responsavel_id
        LEFT JOIN defeitos d ON d.id=o.defeito_id
        LEFT JOIN causas c ON c.id=o.causa_id
        WHERE o.data_abertura::date BETWEEN %s AND %s
        ORDER BY o.numero""", (ini, fim))

    colunas = ["OS", "Abertura", "Estabelecimento", "Tipo", "Setor", "Cód. equip.",
               "Equipamento", "Crit.", "Manutenção", "Origem", "Problema",
               "Solicitante", "Responsável", "Status", "Máq. parada", "Início",
               "Conclusão", "Aprovação", "Defeito", "Causa", "Ação realizada",
               "Tempo (h)", "Custo peças", "Custo MO", "Custo total", "Reprovação"]
    linhas = []
    for o in dados or []:
        horas = round((o["tempo_trabalho_seg"] or 0) / 3600, 2)
        pecas = float(o["custo_pecas"] or 0)
        mo = float(o["custo_hh"] or 0)
        linhas.append([
            o["numero"], o["data_abertura"], o["estab"], o["tipo"], o["setor"],
            o["eq_codigo"], o["equipamento"], o["criticidade"], o["tipo_manutencao"],
            o["origem"], o["descricao_problema"], o["solicitante"], o["responsavel"],
            o["status"], o["maquina_parada"], o["data_inicio"], o["data_conclusao"],
            o["data_aprovacao"], o["defeito"], o["causa"], o["acao_realizada"],
            horas, pecas, mo, round(pecas + mo, 2), o["comentario_reprova"]])

    wb = Workbook()
    escrever_aba(wb.active, "ORDENS DE SERVIÇO — CORRETIVAS", colunas, linhas,
                 rotulo_periodo(ini, fim),
                 larguras=[7, 17, 15, 11, 16, 12, 34, 6, 13, 12, 46, 20, 20, 18, 11,
                           17, 17, 17, 15, 22, 46, 10, 13, 13, 13, 34],
                 formatos={22: "#,##0.00", 23: MOEDA, 24: MOEDA, 25: MOEDA},
                 totalizar=[22, 23, 24, 25])
    wb.active.title = "Ordens de Serviço"

    # ── Resumo por status e por equipamento ──
    resumo = db.query("""SELECT o.status, COUNT(*) AS qt,
                                COALESCE(AVG(o.tempo_trabalho_seg),0)/3600 AS media_h
                         FROM ordens_servico o
                         WHERE o.data_abertura::date BETWEEN %s AND %s
                         GROUP BY o.status ORDER BY qt DESC""", (ini, fim))
    escrever_aba(wb.create_sheet("Resumo por status"), "RESUMO POR SITUAÇÃO",
                 ["Situação", "Quantidade", "Tempo médio (h)"],
                 [[r["status"], r["qt"], round(float(r["media_h"]), 2)] for r in resumo or []],
                 rotulo_periodo(ini, fim), larguras=[26, 14, 18],
                 formatos={3: "#,##0.00"}, totalizar=[2])

    por_eq = db.query("""SELECT e.codigo, e.nome, e.criticidade, COUNT(o.id) AS qt,
                                COALESCE(SUM(o.tempo_trabalho_seg),0)/3600 AS horas,
                                COALESCE(SUM(o.custo_pecas+o.custo_hh),0) AS custo
                         FROM ordens_servico o JOIN equipamentos e ON e.id=o.equipamento_id
                         WHERE o.data_abertura::date BETWEEN %s AND %s
                         GROUP BY e.id ORDER BY qt DESC""", (ini, fim))
    escrever_aba(wb.create_sheet("Por equipamento"), "OCORRÊNCIAS POR EQUIPAMENTO",
                 ["Código", "Equipamento", "Crit.", "Nº de OS", "Horas", "Custo"],
                 [[r["codigo"], r["nome"], r["criticidade"], r["qt"],
                   round(float(r["horas"]), 2), float(r["custo"])] for r in por_eq or []],
                 rotulo_periodo(ini, fim), larguras=[12, 38, 7, 11, 12, 15],
                 formatos={5: "#,##0.00", 6: MOEDA}, totalizar=[4, 5, 6])
    return enviar(wb, "relatorio_ordens_servico")


@bp.route("/os-apontamentos")
@exige("relatorios")
def rel_apontamentos():
    ini, fim = periodo()
    ap = db.query("""SELECT o.numero, a.criado_em, u.nome AS usuario, a.tipo, a.descricao
                     FROM os_apontamentos a
                     JOIN ordens_servico o ON o.id=a.os_id
                     LEFT JOIN usuarios u ON u.id=a.usuario_id
                     WHERE a.criado_em::date BETWEEN %s AND %s
                     ORDER BY o.numero, a.criado_em""", (ini, fim))
    wb = Workbook()
    escrever_aba(wb.active, "APONTAMENTOS DAS ORDENS DE SERVIÇO",
                 ["OS", "Data/Hora", "Usuário", "Tipo", "Descrição"],
                 [[a["numero"], a["criado_em"], a["usuario"], a["tipo"], a["descricao"]]
                  for a in ap or []],
                 rotulo_periodo(ini, fim), larguras=[8, 18, 24, 18, 80])
    wb.active.title = "Apontamentos"

    tempos = db.query("""SELECT o.numero, u.nome AS usuario, t.tipo, t.inicio, t.fim,
                                COALESCE(t.duracao_seg,0)/3600.0 AS horas
                         FROM os_tempos t
                         JOIN ordens_servico o ON o.id=t.os_id
                         LEFT JOIN usuarios u ON u.id=t.usuario_id
                         WHERE t.inicio::date BETWEEN %s AND %s
                         ORDER BY o.numero, t.inicio""", (ini, fim))
    escrever_aba(wb.create_sheet("Tempos"), "APONTAMENTO DE TEMPO",
                 ["OS", "Manutentor", "Tipo", "Início", "Fim", "Duração (h)"],
                 [[t["numero"], t["usuario"], t["tipo"], t["inicio"], t["fim"],
                   round(float(t["horas"]), 3)] for t in tempos or []],
                 rotulo_periodo(ini, fim), larguras=[8, 24, 18, 18, 18, 13],
                 formatos={6: "#,##0.000"}, totalizar=[6])
    return enviar(wb, "relatorio_apontamentos")


# ══════════════════════════════════════════════════════════════════
#  PREVENTIVAS
# ══════════════════════════════════════════════════════════════════
@bp.route("/preventivas")
@exige("relatorios")
def rel_preventivas():
    ano = int(request.args.get("ano") or db.hoje().year)
    wb = Workbook()

    oms = db.query("""SELECT om.numero, e.codigo AS eq_codigo, e.nome AS eq_nome,
                             p.nome AS plano, p.codigo_doc, om.periodicidade,
                             om.ano, om.semana, om.data_prevista, om.data_inicio,
                             om.data_fim, m1.nome AS manutentor1, m2.nome AS manutentor2,
                             om.terceirizado, om.empresa, om.horimetro, om.status,
                             om.no_prazo, l.nome AS visto_lider, om.visto_lider_em,
                             COALESCE(om.tempo_total_seg,0)/3600.0 AS horas, om.observacoes
                      FROM ordens_manutencao om
                      LEFT JOIN equipamentos e ON e.id=om.equipamento_id
                      LEFT JOIN planos_preventiva p ON p.id=om.plano_id
                      LEFT JOIN usuarios m1 ON m1.id=om.manutentor1_id
                      LEFT JOIN usuarios m2 ON m2.id=om.manutentor2_id
                      LEFT JOIN usuarios l ON l.id=om.visto_lider_id
                      WHERE om.ano=%s ORDER BY om.numero""", (ano,))
    escrever_aba(wb.active, f"ORDENS DE MANUTENÇÃO PREVENTIVA — {ano}",
                 ["OM", "Cód. equip.", "Equipamento", "Plano", "Documento", "Período",
                  "Ano", "Semana", "Prevista", "Início", "Fim", "Manutentor 1",
                  "Manutentor 2", "Terceiros", "Empresa", "Horímetro", "Status",
                  "No prazo", "Visto do líder", "Data do visto", "Horas", "Observações"],
                 [[o["numero"], o["eq_codigo"], o["eq_nome"], o["plano"], o["codigo_doc"],
                   o["periodicidade"], o["ano"], o["semana"], o["data_prevista"],
                   o["data_inicio"], o["data_fim"], o["manutentor1"], o["manutentor2"],
                   o["terceirizado"], o["empresa"], o["horimetro"], o["status"],
                   o["no_prazo"], o["visto_lider"], o["visto_lider_em"],
                   round(float(o["horas"]), 2), o["observacoes"]] for o in oms or []],
                 larguras=[7, 12, 32, 30, 12, 9, 7, 9, 13, 17, 17, 20, 20, 10, 20,
                           12, 13, 10, 20, 17, 10, 44],
                 formatos={21: "#,##0.00"}, totalizar=[21])
    wb.active.title = "Ordens de Manutenção"

    # Programação 52 semanas
    prog = db.query("""SELECT e.codigo, e.nome, p.nome AS plano, pr.semana,
                              pr.periodicidade, pr.status, om.numero AS om
                       FROM programacao pr
                       JOIN planos_preventiva p ON p.id=pr.plano_id
                       JOIN equipamentos e ON e.id=pr.equipamento_id
                       LEFT JOIN ordens_manutencao om ON om.id=pr.om_id
                       WHERE pr.ano=%s ORDER BY e.codigo, pr.semana""", (ano,))
    escrever_aba(wb.create_sheet("Programação"), f"PROGRAMAÇÃO DE PREVENTIVAS — {ano}",
                 ["Cód. equip.", "Equipamento", "Plano", "Semana", "Período",
                  "Situação", "OM gerada"],
                 [[p["codigo"], p["nome"], p["plano"], p["semana"], p["periodicidade"],
                   p["status"], p["om"]] for p in prog or []],
                 larguras=[12, 34, 30, 10, 10, 13, 12])

    # Previsto x realizado por mês
    linhas = []
    for m in range(1, 13):
        previsto = db.scalar("""SELECT COUNT(*) AS n FROM programacao
                                WHERE ano=%s AND EXTRACT(MONTH FROM
                                  make_date(ano,1,1)+(semana-1)*7)=%s""", (ano, m), default=0)
        realizado = db.scalar("""SELECT COUNT(*) AS n FROM ordens_manutencao
                                 WHERE status='concluida' AND ano=%s
                                   AND EXTRACT(MONTH FROM data_fim)=%s
                                   AND COALESCE(no_prazo,TRUE)=TRUE""", (ano, m), default=0)
        pct = round(realizado / previsto * 100, 1) if previsto else 0
        linhas.append([["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO",
                        "SET", "OUT", "NOV", "DEZ"][m - 1], previsto, realizado, pct])
    escrever_aba(wb.create_sheet("Previsto x Realizado"),
                 f"ATENDIMENTO DE PREVENTIVAS — {ano}",
                 ["Mês", "Previsto", "Realizado no prazo", "% atendimento"], linhas,
                 larguras=[10, 12, 20, 16], formatos={4: '0.0"%"'}, totalizar=[2, 3])

    # Por responsável
    resp = db.query("""SELECT u.nome, COUNT(om.id) AS total,
                              COUNT(*) FILTER (WHERE om.status='concluida') AS concluidas,
                              COUNT(*) FILTER (WHERE om.status='concluida' AND om.no_prazo) AS no_prazo
                       FROM usuarios u JOIN ordens_manutencao om ON om.manutentor1_id=u.id
                       WHERE om.ano=%s GROUP BY u.id, u.nome ORDER BY total DESC""", (ano,))
    escrever_aba(wb.create_sheet("Por responsável"), f"PREVENTIVAS POR RESPONSÁVEL — {ano}",
                 ["Manutentor", "OMs atribuídas", "Concluídas", "No prazo", "% no prazo"],
                 [[r["nome"], r["total"], r["concluidas"], r["no_prazo"],
                   round(r["no_prazo"] / r["total"] * 100, 1) if r["total"] else 0]
                  for r in resp or []],
                 larguras=[26, 16, 14, 12, 14], formatos={5: '0.0"%"'},
                 totalizar=[2, 3, 4])
    return enviar(wb, f"relatorio_preventivas_{ano}")


@bp.route("/checklists")
@exige("relatorios")
def rel_checklists():
    ano = int(request.args.get("ano") or db.hoje().year)
    resp = db.query("""SELECT om.numero, e.codigo AS eq, p.nome AS plano,
                              om.data_fim, ci.numero AS item, ci.descricao,
                              ci.periodicidade, r.resposta, r.observacao,
                              os.numero AS os_gerada
                       FROM om_respostas r
                       JOIN ordens_manutencao om ON om.id=r.om_id
                       JOIN checklist_itens ci ON ci.id=r.item_id
                       LEFT JOIN equipamentos e ON e.id=om.equipamento_id
                       LEFT JOIN planos_preventiva p ON p.id=om.plano_id
                       LEFT JOIN ordens_servico os ON os.id=r.os_gerada
                       WHERE om.ano=%s ORDER BY om.numero, ci.ordem""", (ano,))
    wb = Workbook()
    escrever_aba(wb.active, f"RESPOSTAS DOS CHECK LISTS — {ano}",
                 ["OM", "Equipamento", "Plano", "Conclusão", "Item", "Descrição",
                  "Período", "Resposta", "Observação", "OS gerada"],
                 [[r["numero"], r["eq"], r["plano"], r["data_fim"], r["item"],
                   r["descricao"], r["periodicidade"], r["resposta"], r["observacao"],
                   r["os_gerada"]] for r in resp or []],
                 larguras=[7, 13, 30, 17, 8, 50, 9, 11, 40, 11])
    wb.active.title = "Check lists"
    return enviar(wb, f"relatorio_checklists_{ano}")


# ══════════════════════════════════════════════════════════════════
#  RONDAS
# ══════════════════════════════════════════════════════════════════
@bp.route("/rondas")
@exige("relatorios")
def rel_rondas():
    ini, fim = periodo()
    dados = db.query("""SELECT ex.data, r.nome AS ronda, u.nome AS usuario, ex.status,
                               p.ordem, p.descricao, e.codigo AS eq,
                               rr.resposta, rr.observacao, os.numero AS os_gerada
                        FROM ronda_respostas rr
                        JOIN ronda_execucoes ex ON ex.id=rr.execucao_id
                        JOIN rondas r ON r.id=ex.ronda_id
                        JOIN ronda_pontos p ON p.id=rr.ponto_id
                        LEFT JOIN equipamentos e ON e.id=p.equipamento_id
                        LEFT JOIN usuarios u ON u.id=ex.usuario_id
                        LEFT JOIN ordens_servico os ON os.id=rr.os_gerada
                        WHERE ex.data BETWEEN %s AND %s
                        ORDER BY ex.data DESC, p.ordem""", (ini, fim))
    wb = Workbook()
    escrever_aba(wb.active, "RONDAS DIÁRIAS DE INSPEÇÃO",
                 ["Data", "Ronda", "Executada por", "Situação", "Ponto", "Descrição",
                  "Equipamento", "Resposta", "Observação", "OS gerada"],
                 [[d["data"], d["ronda"], d["usuario"], d["status"], d["ordem"],
                   d["descricao"], d["eq"], d["resposta"], d["observacao"],
                   d["os_gerada"]] for d in dados or []],
                 rotulo_periodo(ini, fim),
                 larguras=[12, 26, 22, 14, 8, 50, 13, 11, 40, 11])
    wb.active.title = "Rondas"
    return enviar(wb, "relatorio_rondas")


# ══════════════════════════════════════════════════════════════════
#  MATERIAIS
# ══════════════════════════════════════════════════════════════════
@bp.route("/materiais")
@exige("relatorios")
def rel_materiais():
    wb = Workbook()
    mats = db.query("""SELECT m.codigo, m.descricao, m.unidade, m.tipo, m.aplicacao,
                              m.localizacao, m.critico, m.estoque_min, m.estoque_max,
                              m.saldo_sap, m.valor_unit, m.ativo, m.atualizado_em,
                 COALESCE(SUM(CASE WHEN mv.tipo IN ('ENTRADA','AJUSTE') THEN mv.quantidade
                                   ELSE -mv.quantidade END),0) AS saldo_nlag
                       FROM materiais m LEFT JOIN movimentacoes mv ON mv.codigo=m.codigo
                       GROUP BY m.id ORDER BY m.codigo""")
    linhas = []
    for m in mats or []:
        saldo = float(m["saldo_nlag"]) if m["tipo"] == "NLAG" else float(m["saldo_sap"] or 0)
        minimo = float(m["estoque_min"] or 0)
        valor = float(m["valor_unit"] or 0)
        linhas.append([
            m["codigo"], m["descricao"], m["unidade"], m["tipo"], m["aplicacao"],
            m["localizacao"], m["critico"], saldo, minimo, float(m["estoque_max"] or 0),
            "ABAIXO DO MÍNIMO" if minimo and saldo < minimo else "OK",
            valor, round(saldo * valor, 2), m["ativo"], m["atualizado_em"]])
    escrever_aba(wb.active, "SALDO DE MATERIAIS",
                 ["Código", "Descrição", "UMB", "Depósito", "Aplicação", "Localização",
                  "Crítico", "Saldo", "Mínimo", "Máximo", "Situação", "Valor unit.",
                  "Valor em estoque", "Ativo", "Atualizado em"],
                 linhas,
                 larguras=[12, 46, 7, 11, 26, 16, 9, 12, 11, 11, 20, 13, 17, 8, 17],
                 formatos={8: NUM, 9: NUM, 10: NUM, 12: MOEDA, 13: MOEDA},
                 totalizar=[13])
    wb.active.title = "Saldo"

    ini, fim = periodo()
    mov = db.query("""SELECT mv.data_hora, mv.codigo, m.descricao, m.unidade, mv.tipo,
                             mv.quantidade, mv.usuario, o.numero AS os, mv.observacao
                      FROM movimentacoes mv
                      LEFT JOIN materiais m ON m.codigo=mv.codigo
                      LEFT JOIN ordens_servico o ON o.id=mv.os_id
                      WHERE mv.data_hora::date BETWEEN %s AND %s
                      ORDER BY mv.data_hora DESC""", (ini, fim))
    escrever_aba(wb.create_sheet("Movimentações"), "MOVIMENTAÇÕES DO DEPÓSITO NLAG",
                 ["Data/Hora", "Código", "Descrição", "UMB", "Tipo", "Quantidade",
                  "Usuário", "OS", "Observação"],
                 [[m["data_hora"], m["codigo"], m["descricao"], m["unidade"], m["tipo"],
                   float(m["quantidade"]), m["usuario"], m["os"], m["observacao"]]
                  for m in mov or []],
                 rotulo_periodo(ini, fim),
                 larguras=[18, 12, 40, 7, 11, 13, 22, 8, 40], formatos={6: NUM})

    # Alertas de estoque mínimo
    alertas = []
    for m in mats or []:
        saldo = float(m["saldo_nlag"]) if m["tipo"] == "NLAG" else float(m["saldo_sap"] or 0)
        minimo = float(m["estoque_min"] or 0)
        if not minimo or saldo >= minimo:
            continue
        maximo = float(m["estoque_max"] or 0)
        consumo = float(db.scalar("""SELECT COALESCE(SUM(quantidade),0) AS q
                                     FROM movimentacoes WHERE codigo=%s AND tipo='SAIDA'
                                       AND data_hora > NOW()-INTERVAL '30 days'""",
                                  (m["codigo"],), default=0))
        sugestao = round(max(maximo - saldo, minimo - saldo), 2)
        alertas.append([m["codigo"], m["descricao"], m["unidade"], m["tipo"], m["critico"],
                        saldo, minimo, maximo, consumo, sugestao,
                        round(sugestao * float(m["valor_unit"] or 0), 2)])
    escrever_aba(wb.create_sheet("Alertas de compra"),
                 "ITENS ABAIXO DO ESTOQUE MÍNIMO",
                 ["Código", "Descrição", "UMB", "Depósito", "Crítico", "Saldo",
                  "Mínimo", "Máximo", "Consumo 30d", "Sugestão de compra", "Custo estimado"],
                 alertas, larguras=[12, 46, 7, 11, 9, 11, 11, 11, 14, 19, 16],
                 formatos={6: NUM, 7: NUM, 8: NUM, 9: NUM, 10: NUM, 11: MOEDA},
                 totalizar=[10, 11])
    return enviar(wb, "relatorio_materiais")


@bp.route("/solicitacoes")
@exige("relatorios")
def rel_solicitacoes():
    ini, fim = periodo()
    dados = db.query("""SELECT s.numero, s.criado_em, u.nome AS solicitante, s.codigo,
                               s.descricao, s.link, s.tipo, s.quantidade,
                               cc.codigo AS cc, cc.nome AS cc_nome, o.numero AS os,
                               s.num_ficha, s.id_4mdg, s.num_pr, s.codigo_final,
                               s.tipo_material, s.dt_solicitacao, s.dt_cadastro,
                               s.dt_chegada, s.situacao, s.observacoes,
                               (CURRENT_DATE - s.criado_em::date) AS dias
                        FROM solicitacoes_material s
                        LEFT JOIN usuarios u ON u.id=s.solicitante_id
                        LEFT JOIN centros_custo cc ON cc.id=s.centro_custo_id
                        LEFT JOIN ordens_servico o ON o.id=s.os_id
                        WHERE s.criado_em::date BETWEEN %s AND %s
                        ORDER BY s.numero""", (ini, fim))
    wb = Workbook()
    escrever_aba(wb.active, "SOLICITAÇÕES DE MATERIAL",
                 ["SM", "Criada em", "Solicitante", "Código", "Descrição", "Link",
                  "Tipo", "Qtd", "C. custo", "Centro de custo", "OS", "Ficha/FDS",
                  "ID 4MDG", "Nº PR", "Código final", "Depósito", "Dt. solicitação",
                  "Dt. cadastro", "Dt. chegada", "Situação", "Observações", "Dias em aberto"],
                 [[s["numero"], s["criado_em"], s["solicitante"], s["codigo"],
                   s["descricao"], s["link"], s["tipo"], float(s["quantidade"]),
                   s["cc"], s["cc_nome"], s["os"], s["num_ficha"], s["id_4mdg"],
                   s["num_pr"], s["codigo_final"], s["tipo_material"],
                   s["dt_solicitacao"], s["dt_cadastro"], s["dt_chegada"],
                   s["situacao"], s["observacoes"], s["dias"]] for s in dados or []],
                 rotulo_periodo(ini, fim),
                 larguras=[7, 17, 22, 12, 46, 30, 15, 9, 12, 24, 8, 12, 12, 13, 13,
                           11, 15, 14, 14, 18, 40, 13],
                 formatos={8: NUM})
    wb.active.title = "Solicitações"

    hist = db.query("""SELECT s.numero, h.criado_em, u.nome AS usuario, h.situacao,
                              h.comentario
                       FROM solicitacao_historico h
                       JOIN solicitacoes_material s ON s.id=h.solicitacao_id
                       LEFT JOIN usuarios u ON u.id=h.usuario_id
                       WHERE h.criado_em::date BETWEEN %s AND %s
                       ORDER BY s.numero, h.criado_em""", (ini, fim))
    escrever_aba(wb.create_sheet("Histórico"), "HISTÓRICO DAS SOLICITAÇÕES",
                 ["SM", "Data/Hora", "Usuário", "Situação", "Comentário"],
                 [[h["numero"], h["criado_em"], h["usuario"], h["situacao"],
                   h["comentario"]] for h in hist or []],
                 rotulo_periodo(ini, fim), larguras=[7, 18, 24, 20, 60])
    return enviar(wb, "relatorio_solicitacoes")


# ══════════════════════════════════════════════════════════════════
#  EQUIPAMENTOS E INDICADORES
# ══════════════════════════════════════════════════════════════════
@bp.route("/equipamentos")
@exige("relatorios")
def rel_equipamentos():
    eq = db.query("""SELECT e.*, ct.nome AS setor, est.nome AS estab,
                            cr.nome AS crit_nome, u.nome AS avaliador
                     FROM equipamentos e
                     LEFT JOIN centros_trabalho ct ON ct.id=e.centro_trabalho_id
                     LEFT JOIN estabelecimentos est ON est.id=e.estabelecimento_id
                     LEFT JOIN criticidades cr ON cr.codigo=e.criticidade
                     LEFT JOIN usuarios u ON u.id=e.mtz_avaliado_por
                     ORDER BY e.codigo""")
    wb = Workbook()
    escrever_aba(wb.active, "INVENTÁRIO DE EQUIPAMENTOS",
                 ["Código", "Grupo", "Sub", "Equipamento", "Setor", "Estabelecimento",
                  "Tipo", "Criticidade", "Nível", "Fabricante", "Nº série", "Patrimônio",
                  "Ano", "Capacidade", "Horímetro", "Custo/h parada", "Status", "Ativo"],
                 [[e["codigo"], e["grupo_prev"], e["subcodigo"], e["nome"], e["setor"],
                   e["estab"], e["tipo"], e["criticidade"], e["crit_nome"],
                   e["fabricante"], e["n_serie"], e["patrimonio"], e["ano_fabricacao"],
                   e["capacidade"], float(e["horimetro"] or 0),
                   float(e["custo_hora_parada"] or 0), e["status"], e["ativo"]]
                  for e in eq or []],
                 larguras=[12, 9, 6, 40, 20, 16, 12, 12, 14, 18, 18, 14, 8, 18,
                           12, 15, 12, 8],
                 formatos={15: NUM, 16: MOEDA})
    wb.active.title = "Inventário"

    escrever_aba(wb.create_sheet("Matriz de criticidade"),
                 "MATRIZ DE CLASSIFICAÇÃO DE CRITICIDADE",
                 ["Código", "Equipamento", "Segurança", "Produção", "Qualidade",
                  "Frequência", "Reparo", "Redundância", "Pontuação", "Criticidade",
                  "Avaliado em", "Avaliado por", "Justificativa"],
                 [[e["codigo"], e["nome"], e["mtz_seguranca"], e["mtz_producao"],
                   e["mtz_qualidade"], e["mtz_frequencia"], e["mtz_reparo"],
                   e["mtz_redundancia"],
                   float(e["mtz_pontuacao"]) if e["mtz_pontuacao"] is not None else None,
                   e["criticidade"], e["mtz_avaliado_em"], e["avaliador"],
                   e["mtz_justificativa"]] for e in eq or []],
                 larguras=[12, 40, 11, 11, 11, 12, 10, 13, 12, 12, 17, 22, 44],
                 formatos={9: "#,##0.0"})

    niveis = db.query("SELECT * FROM criticidades ORDER BY ordem")
    escrever_aba(wb.create_sheet("Níveis"), "NÍVEIS DE CRITICIDADE",
                 ["Nível", "Nome", "Ordem", "Resposta (h)", "Conclusão (h)",
                  "Ativo", "Quando usar"],
                 [[n["codigo"], n["nome"], n["ordem"],
                   float(n["sla_resposta_h"]) if n["sla_resposta_h"] is not None else None,
                   float(n["sla_conclusao_h"]) if n["sla_conclusao_h"] is not None else None,
                   n["ativo"], n["descricao"]] for n in niveis or []],
                 larguras=[8, 18, 8, 14, 15, 8, 70])
    return enviar(wb, "relatorio_equipamentos")


@bp.route("/indicadores")
@exige("relatorios")
def rel_indicadores():
    ini, fim = periodo()
    dias = max((date.fromisoformat(fim) - date.fromisoformat(ini)).days, 1)
    horas_periodo = dias * 24

    eq = db.query("""SELECT e.codigo, e.nome, e.criticidade, e.status,
                            COUNT(o.id) AS falhas,
                            COALESCE(SUM(o.tempo_trabalho_seg),0) AS reparo,
                            COALESCE(SUM(o.custo_pecas),0) AS pecas,
                            COALESCE(SUM(o.custo_hh),0) AS mo
                     FROM equipamentos e
                     LEFT JOIN ordens_servico o ON o.equipamento_id=e.id
                          AND o.tipo_manutencao IN ('corretiva','intervencao')
                          AND o.status='concluida'
                          AND o.data_conclusao::date BETWEEN %s AND %s
                     WHERE e.ativo=TRUE GROUP BY e.id ORDER BY falhas DESC, e.codigo""",
                  (ini, fim))
    linhas = []
    for e in eq or []:
        f = e["falhas"] or 0
        rep_h = (e["reparo"] or 0) / 3600
        pecas, mo = float(e["pecas"] or 0), float(e["mo"] or 0)
        linhas.append([
            e["codigo"], e["nome"], e["criticidade"], e["status"], f,
            round(rep_h, 2),
            round(rep_h / f, 2) if f else 0,
            round((horas_periodo - rep_h) / f, 1) if f else 0,
            round((horas_periodo - rep_h) / horas_periodo * 100, 2),
            pecas, mo, round(pecas + mo, 2)])
    wb = Workbook()
    escrever_aba(wb.active, "INDICADORES DE MANUTENÇÃO POR EQUIPAMENTO",
                 ["Código", "Equipamento", "Crit.", "Status", "Nº de falhas",
                  "Horas de reparo", "MTTR (h)", "MTBF (h)", "Disponibilidade %",
                  "Custo peças", "Custo MO", "Custo total"],
                 linhas, rotulo_periodo(ini, fim),
                 larguras=[12, 40, 7, 12, 13, 16, 11, 12, 17, 14, 14, 14],
                 formatos={6: "#,##0.00", 7: "#,##0.00", 8: "#,##0.0",
                           9: "#,##0.00", 10: MOEDA, 11: MOEDA, 12: MOEDA},
                 totalizar=[5, 6, 10, 11, 12])
    wb.active.title = "Por equipamento"

    resp = db.query("""SELECT u.nome, COUNT(o.id) AS total,
                              COUNT(*) FILTER (WHERE o.status='concluida') AS concluidas,
                              COALESCE(SUM(o.tempo_trabalho_seg),0)/3600.0 AS horas,
                       COALESCE(AVG(o.tempo_trabalho_seg) FILTER (WHERE o.status='concluida'),0)/3600.0 AS media
                       FROM usuarios u JOIN ordens_servico o ON o.responsavel_id=u.id
                       WHERE o.data_abertura::date BETWEEN %s AND %s
                       GROUP BY u.id, u.nome ORDER BY total DESC""", (ini, fim))
    escrever_aba(wb.create_sheet("Por responsável"), "PRODUTIVIDADE POR MANUTENTOR",
                 ["Manutentor", "OS atribuídas", "Concluídas", "Horas trabalhadas",
                  "Tempo médio (h)"],
                 [[r["nome"], r["total"], r["concluidas"], round(float(r["horas"]), 2),
                   round(float(r["media"]), 2)] for r in resp or []],
                 rotulo_periodo(ini, fim), larguras=[26, 15, 13, 18, 16],
                 formatos={4: "#,##0.00", 5: "#,##0.00"}, totalizar=[2, 3, 4])

    dc = db.query("""SELECT d.nome AS defeito, c.nome AS causa, COUNT(*) AS qt,
                            COALESCE(SUM(o.tempo_trabalho_seg),0)/3600.0 AS horas
                     FROM ordens_servico o
                     LEFT JOIN defeitos d ON d.id=o.defeito_id
                     LEFT JOIN causas c ON c.id=o.causa_id
                     WHERE o.status='concluida' AND o.data_conclusao::date BETWEEN %s AND %s
                     GROUP BY d.nome, c.nome ORDER BY qt DESC""", (ini, fim))
    escrever_aba(wb.create_sheet("Defeitos e causas"), "DEFEITOS E CAUSAS",
                 ["Tipo de defeito", "Causa", "Ocorrências", "Horas"],
                 [[d["defeito"], d["causa"], d["qt"], round(float(d["horas"]), 2)]
                  for d in dc or []],
                 rotulo_periodo(ini, fim), larguras=[24, 34, 14, 12],
                 formatos={4: "#,##0.00"}, totalizar=[3, 4])
    return enviar(wb, "relatorio_indicadores")


@bp.route("/terceiros")
@exige("relatorios")
def rel_terceiros():
    t = db.query("""SELECT t.data_envio, e.codigo AS eq, e.nome AS eq_nome, t.empresa,
                           t.tipo_servico, t.descricao, t.data_retorno, t.valor,
                           t.recebido_por, u.nome AS manutentor
                    FROM manutencoes_terceiros t
                    LEFT JOIN equipamentos e ON e.id=t.equipamento_id
                    LEFT JOIN usuarios u ON u.id=t.manutentor_id
                    ORDER BY t.data_envio DESC NULLS LAST""")
    wb = Workbook()
    escrever_aba(wb.active, "MANUTENÇÕES EM TERCEIROS",
                 ["Envio", "Cód. equip.", "Equipamento", "Empresa", "Tipo de serviço",
                  "Descrição", "Retorno", "Valor", "Recebido por", "Responsável"],
                 [[x["data_envio"], x["eq"], x["eq_nome"], x["empresa"],
                   x["tipo_servico"], x["descricao"], x["data_retorno"],
                   float(x["valor"] or 0), x["recebido_por"], x["manutentor"]]
                  for x in t or []],
                 larguras=[13, 12, 34, 26, 16, 46, 13, 14, 22, 22],
                 formatos={8: MOEDA}, totalizar=[8])
    wb.active.title = "Terceiros"
    return enviar(wb, "relatorio_terceiros")


@bp.route("/usuarios")
@exige("admin")
def rel_usuarios():
    u = db.query("SELECT * FROM usuarios ORDER BY nome")
    wb = Workbook()
    escrever_aba(wb.active, "USUÁRIOS DO SISTEMA",
                 ["Usuário", "Nome", "Matrícula", "E-mail", "Telefone", "Perfil",
                  "Ativo", "Criado em", "Último acesso"],
                 [[x["usuario"], x["nome"], x["matricula"], x["email"], x["telefone"],
                   x["perfil"], x["ativo"], x["criado_em"], x["ultimo_acesso"]]
                  for x in u or []],
                 larguras=[18, 30, 14, 30, 16, 16, 8, 17, 17])
    wb.active.title = "Usuários"

    log = db.query("SELECT * FROM log_auditoria ORDER BY criado_em DESC LIMIT 5000")
    escrever_aba(wb.create_sheet("Auditoria"), "LOG DE AUDITORIA",
                 ["Data/Hora", "Usuário", "Ação", "Entidade", "ID", "Detalhe"],
                 [[l["criado_em"], l["usuario"], l["acao"], l["entidade"],
                   l["entidade_id"], l["detalhe"]] for l in log or []],
                 larguras=[18, 22, 24, 20, 8, 50])
    return enviar(wb, "relatorio_usuarios")


# ══════════════════════════════════════════════════════════════════
#  BACKUP
# ══════════════════════════════════════════════════════════════════
COLUNAS_BINARIAS = {"imagem", "dados", "foto", "senha_hash"}


def tabelas_do_banco():
    t = db.query("""SELECT table_name FROM information_schema.tables
                    WHERE table_schema='public' AND table_type='BASE TABLE'
                    ORDER BY table_name""")
    return [x["table_name"] for x in t or []]


def colunas_exportaveis(tabela):
    cols = db.query("""SELECT column_name, data_type FROM information_schema.columns
                       WHERE table_schema='public' AND table_name=%s
                       ORDER BY ordinal_position""", (tabela,))
    saida = []
    for c in cols or []:
        if c["column_name"] in COLUNAS_BINARIAS or c["data_type"] == "bytea":
            continue
        saida.append(c["column_name"])
    return saida


@bp.route("/backup")
@exige("backup")
def backup():
    tabelas = []
    total = 0
    for t in tabelas_do_banco():
        n = db.scalar(f'SELECT COUNT(*) AS n FROM "{t}"', default=0)
        total += n
        tabelas.append({"nome": t, "registros": n,
                        "colunas": len(colunas_exportaveis(t))})
    tabelas.sort(key=lambda x: -x["registros"])

    anexos = {
        "os": db.scalar("SELECT COUNT(*) AS n FROM os_anexos", default=0),
        "om": db.scalar("SELECT COUNT(*) AS n FROM om_anexos", default=0),
        "fotos_ronda": db.scalar(
            "SELECT COUNT(*) AS n FROM ronda_respostas WHERE foto IS NOT NULL", default=0),
        "imagens_material": db.scalar(
            "SELECT COUNT(*) AS n FROM materiais WHERE imagem IS NOT NULL", default=0),
    }
    ultimo = db.um("""SELECT criado_em, usuario FROM log_auditoria
                      WHERE acao='backup' ORDER BY id DESC LIMIT 1""")
    return render_template("rel/backup.html", tabelas=tabelas, total=total,
                           anexos=anexos, ultimo=ultimo)


@bp.route("/backup/excel")
@exige("backup")
def backup_excel():
    """Backup completo em Excel — uma aba por tabela."""
    wb = Workbook()
    wb.remove(wb.active)

    resumo = []
    for tabela in tabelas_do_banco():
        cols = colunas_exportaveis(tabela)
        if not cols:
            continue
        lista = ", ".join(f'"{c}"' for c in cols)
        dados = db.query(f'SELECT {lista} FROM "{tabela}" ORDER BY 1')
        resumo.append([tabela, len(dados or []), len(cols)])
        # O Excel limita o nome da aba a 31 caracteres
        ws = wb.create_sheet(tabela[:31])
        escrever_aba(ws, f"TABELA: {tabela.upper()}", cols,
                     [[r[c] for c in cols] for r in dados or []],
                     f"Backup gerado em {db.agora():%d/%m/%Y %H:%M} "
                     f"por {session.get('nome', '')}")

    indice = wb.create_sheet("_ÍNDICE", 0)
    escrever_aba(indice, "BACKUP COMPLETO — SISTEMA DE MANUTENÇÃO",
                 ["Tabela", "Registros", "Colunas"], resumo,
                 f"Décio Metalúrgica · {db.agora():%d/%m/%Y %H:%M} · "
                 f"{sum(r[1] for r in resumo)} registros · "
                 "Imagens e anexos não estão incluídos (use o backup do PostgreSQL)",
                 larguras=[32, 14, 12], totalizar=[2])

    db.registrar_log(session["uid"], session["nome"], "backup", detalhe="Excel completo")
    return enviar(wb, "backup_completo")


@bp.route("/backup/json")
@exige("backup")
def backup_json():
    """Backup completo em JSON — fidelidade total dos dados."""
    saida = {
        "sistema": "Sistema Centralizado de Manutenção — Décio Metalúrgica",
        "gerado_em": db.agora().isoformat(),
        "gerado_por": session.get("nome"),
        "observacao": ("Colunas binárias (imagens, anexos, fotos) e hashes de senha "
                       "não são exportados. Para um backup integral use pg_dump."),
        "tabelas": {},
    }
    for tabela in tabelas_do_banco():
        cols = colunas_exportaveis(tabela)
        if not cols:
            continue
        lista = ", ".join(f'"{c}"' for c in cols)
        dados = db.query(f'SELECT {lista} FROM "{tabela}" ORDER BY 1')
        saida["tabelas"][tabela] = [
            {c: _json_valor(r[c]) for c in cols} for r in dados or []]

    conteudo = json.dumps(saida, ensure_ascii=False, indent=1, default=str)
    db.registrar_log(session["uid"], session["nome"], "backup", detalhe="JSON completo")
    arquivo = f"backup_completo_{db.agora():%Y%m%d_%H%M}.json"
    return Response(conteudo.encode("utf-8"), mimetype="application/json",
                    headers={"Content-Disposition": f"attachment; filename={arquivo}"})


def _json_valor(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (bytes, memoryview)):
        return None
    return v
