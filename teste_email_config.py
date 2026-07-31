"""Testa a configuração de e-mail pela tela, a criptografia da senha e os links."""
import io
import json
import os

os.environ.setdefault("DATABASE_URL", "postgresql://postgres:teste@localhost:5432/manutencao")
os.environ.setdefault("SECRET_KEY", "chave-de-teste-do-sistema")
# Sem variáveis de SMTP no ambiente — tudo deve vir da tela
for v in ("SMTP_HOST", "SMTP_USUARIO", "SMTP_SENHA", "SMTP_REMETENTE", "APP_URL"):
    os.environ.pop(v, None)

import email_config
import mailer
from app import app
from werkzeug.security import generate_password_hash
import db

adm = app.test_client()
adm.post("/login", data={"usuario": "admin", "senha": "teste123"})

print("── Antes de configurar ──")
db.executar("DELETE FROM parametros WHERE chave LIKE 'smtp%%' OR chave='app_url'")
assert mailer.configurado() is False
print("   ✅ sistema reconhece que não há envio configurado")
r = adm.get("/admin/email")
assert r.status_code == 200 and "Falta configurar".encode() in r.data
print("   ✅ tela avisa que falta configurar")

ok, erro = mailer.enviar_agora(["x@y.com"], "teste", "<p>x</p>")
assert ok is False and "não configurado" in erro.lower()
print("   ✅ envio recusado com mensagem clara")

# ══ CONFIGURAR PELA TELA ══════════════════════════════════════
print("\n── Configurando o Gmail pela tela ──")
r = adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "gmail",
    "smtp_usuario": "engenhariadecio2026@gmail.com",
    "smtp_senha": "abcd efgh ijkl mnop",   # com espaços, como o Google entrega
    "smtp_nome_remetente": "Manutenção — Décio Metalúrgica",
    "app_url": "", "smtp_ativo": "1"}, follow_redirects=True)
assert r.status_code == 200

c = email_config.configuracao()
print(f"   host: {c['smtp_host']}:{c['smtp_porta']} ({c['smtp_seguranca']})")
print(f"   conta: {c['smtp_usuario']}")
print(f"   remetente: {c['smtp_remetente']}")
assert c["smtp_host"] == "smtp.gmail.com" and c["smtp_porta"] == 587
assert c["smtp_seguranca"] == "tls"
print("   ✅ escolher 'Gmail' preencheu servidor, porta e segurança")
assert c["smtp_remetente"] == "engenhariadecio2026@gmail.com"
print("   ✅ remetente assumiu a conta informada")
assert mailer.configurado() is True
print("   ✅ envio passou a estar habilitado")

# ══ SENHA ═════════════════════════════════════════════════════
print("\n── Senha ──")
assert email_config.ler_senha() == "abcdefghijklmnop"
print("   ✅ espaços removidos automaticamente")

guardada = db.scalar("SELECT valor FROM parametros WHERE chave='smtp_senha'", default="")
assert "abcdefghijklmnop" not in guardada and len(guardada) > 40
print(f"   ✅ gravada criptografada: {guardada[:34]}…")

r = adm.get("/admin/email")
assert b"abcdefghijklmnop" not in r.data and b"abcd efgh" not in r.data
print("   ✅ nunca aparece na tela")

# Salvar sem informar a senha mantém a anterior
adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "gmail",
    "smtp_usuario": "engenhariadecio2026@gmail.com", "smtp_senha": "",
    "smtp_nome_remetente": "Manutenção", "smtp_ativo": "1"}, follow_redirects=True)
assert email_config.ler_senha() == "abcdefghijklmnop"
print("   ✅ salvar com o campo em branco mantém a senha")

# Backup não leva a senha
r = adm.get("/relatorios/backup/json")
dados = json.loads(r.data.decode("utf-8"))
chaves = [p["chave"] for p in dados["tabelas"].get("parametros", [])]
assert "smtp_senha" not in chaves, "a senha não pode entrar no backup"
assert "smtp_host" in chaves, "as demais configurações devem entrar"
print("   ✅ backup JSON traz a configuração, mas não a senha")

from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(adm.get("/relatorios/backup/excel").data))
ws = wb["parametros"]
valores = [ws.cell(row=i, column=1).value for i in range(5, ws.max_row + 1)]
assert "smtp_senha" not in valores
print("   ✅ backup Excel idem")

# ══ ENDEREÇO AUTOMÁTICO ═══════════════════════════════════════
print("\n── Endereço do sistema ──")
with app.test_request_context("/", base_url="https://manutencao.up.railway.app"):
    detectado = mailer.base_url()
    print(f"   detectado sozinho: {detectado}")
    assert detectado == "https://manutencao.up.railway.app"
print("   ✅ deduz o endereço do próprio acesso quando não é informado")

adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "gmail",
    "smtp_usuario": "engenhariadecio2026@gmail.com",
    "app_url": "https://manutencao.decio.com.br/", "smtp_ativo": "1"},
    follow_redirects=True)
assert mailer.base_url() == "https://manutencao.decio.com.br"
print("   ✅ endereço informado tem prioridade (e a barra final é removida)")

# ══ LINKS DIRETOS ═════════════════════════════════════════════
print("\n── Links dos e-mails ──")
mailer.modo_teste(True)
caixa = mailer.caixa_de_teste()

for usuario, nome, perfil, mail in [
        ("charles", "Charles Pfleger", "solicitante", "ch1000328@intelbras.com.br"),
        ("jaime", "Jaime Matias", "manutentor", "ja1001070@intelbras.com.br"),
        ("lourivaldo", "Lourivaldo Vieira", "lider", "lo1000673@intelbras.com.br")]:
    db.executar("""INSERT INTO usuarios (usuario, senha_hash, nome, perfil, email)
                   VALUES (%s,%s,%s,%s,%s) ON CONFLICT (usuario) DO UPDATE
                   SET perfil=EXCLUDED.perfil, nome=EXCLUDED.nome, email=EXCLUDED.email""",
                (usuario, generate_password_hash("teste123"), nome, perfil, mail))

def entrar(u):
    cli = app.test_client(); cli.post("/login", data={"usuario": u, "senha": "teste123"})
    return cli

sol, lid, mnt = entrar("charles"), entrar("lourivaldo"), entrar("jaime")
est = db.um("SELECT id FROM estabelecimentos WHERE nome='Matriz'")
eq = db.um("SELECT id FROM equipamentos WHERE codigo='GU01-00'")

caixa.clear()
sol.post("/os/nova", data={"estabelecimento_id": est["id"], "tipo": "Industrial",
                           "equipamento_id": eq["id"],
                           "descricao_problema": "Guilhotina desregulada"},
         follow_redirects=True)
o = db.um("SELECT * FROM ordens_servico ORDER BY id DESC LIMIT 1")
assert "https://manutencao.decio.com.br/os/triagem" in caixa[-1]["html"]
print("   ✅ nova OS → link para a tela de triagem")

caixa.clear()
id_jaime = db.scalar("SELECT id FROM usuarios WHERE usuario='jaime'")
lid.post(f"/os/{o['id']}/assumir", data={"responsavel_id": id_jaime}, follow_redirects=True)
assert f"/os/{o['id']}#cronometro" in caixa[-1]["html"]
print("   ✅ OS atribuída → link direto no cronômetro")

mnt.post(f"/os/{o['id']}/acao/iniciar", follow_redirects=True)
caixa.clear()
d = db.um("SELECT id FROM defeitos LIMIT 1")
ca = db.um("SELECT id FROM causas LIMIT 1")
mnt.post(f"/os/{o['id']}/concluir",
         data={"defeito_id": d["id"], "causa_id": ca["id"],
               "acao_realizada": "Regulagem refeita"}, follow_redirects=True)
assert f"/os/{o['id']}#aprovacao" in caixa[-1]["html"]
print("   ✅ OS concluída → link direto no bloco de aprovação")

# As âncoras existem mesmo no HTML das telas
html = mnt.get(f"/os/{o['id']}").data.decode()
assert 'id="cronometro"' in html or 'id="aprovacao"' in html
html_sol = sol.get(f"/os/{o['id']}").data.decode()
assert 'id="aprovacao"' in html_sol
print("   ✅ as âncoras existem nas telas — o link não cai no vazio")

# ══ DESATIVAR ═════════════════════════════════════════════════
print("\n── Desativar o envio ──")
mailer.modo_teste(False)
adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "gmail",
    "smtp_usuario": "engenhariadecio2026@gmail.com",
    "app_url": "https://manutencao.decio.com.br"}, follow_redirects=True)
assert mailer.configurado() is False
print("   ✅ desmarcar 'envio ativo' interrompe os disparos")
r = adm.get("/admin/email")
assert "desativado".encode() in r.data
print("   ✅ tela mostra o estado desativado")

adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "gmail",
    "smtp_usuario": "engenhariadecio2026@gmail.com",
    "app_url": "https://manutencao.decio.com.br", "smtp_ativo": "1"},
    follow_redirects=True)

# ══ SECRET_KEY TROCADA ════════════════════════════════════════
print("\n── SECRET_KEY trocada ──")
original = os.environ["SECRET_KEY"]
os.environ["SECRET_KEY"] = "uma-chave-completamente-diferente"
assert email_config.ler_senha() == ""
assert email_config.senha_ilegivel() is True
print("   ✅ senha antiga fica ilegível, sem quebrar o sistema")
r = adm.get("/admin/email")
assert "não pode mais ser lida".encode() in r.data
print("   ✅ tela orienta a cadastrar a senha de novo")
os.environ["SECRET_KEY"] = original
assert email_config.ler_senha() == "abcdefghijklmnop"
print("   ✅ voltando a chave, a senha volta a ser lida")

# ══ OUTRO PROVEDOR ════════════════════════════════════════════
print("\n── Servidor personalizado ──")
adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "outro",
    "smtp_host": "smtp.empresa.com.br", "smtp_porta": "465",
    "smtp_seguranca": "ssl", "smtp_usuario": "manutencao@decio.com.br",
    "smtp_remetente": "naoresponda@decio.com.br",
    "smtp_senha": "senha123", "smtp_ativo": "1"}, follow_redirects=True)
c = email_config.configuracao()
assert c["smtp_host"] == "smtp.empresa.com.br" and c["smtp_porta"] == 465
assert c["smtp_seguranca"] == "ssl"
assert c["smtp_remetente"] == "naoresponda@decio.com.br"
print("   ✅ aceita servidor, porta, SSL e remetente diferentes da conta")

# ══ PROVEDORES SEM 2FA ════════════════════════════════════════
print("\n── Provedores que dispensam verificação em duas etapas ──")
sem_2fa = [k for k, v in email_config.PROVEDORES.items() if v["sem_2fa"]]
print(f"   {', '.join(sem_2fa)}")
assert "brevo" in sem_2fa and "sendgrid" in sem_2fa
assert email_config.PROVEDORES["gmail"]["sem_2fa"] is False
print("   ✅ Brevo e SendGrid marcados como sem 2FA; Gmail não")

adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "brevo",
    "smtp_usuario": "8f2a10001@smtp-brevo.com",
    "smtp_senha": "xsmtpsib-CHAVE-DE-TESTE",
    "smtp_remetente": "engenhariadecio2026@gmail.com",
    "smtp_nome_remetente": "Manutenção — Décio Metalúrgica",
    "smtp_ativo": "1"}, follow_redirects=True)
c = email_config.configuracao()
print(f"   host: {c['smtp_host']}:{c['smtp_porta']}")
print(f"   login: {c['smtp_usuario']}")
print(f"   remetente: {c['smtp_remetente']}")
assert c["smtp_host"] == "smtp-relay.brevo.com" and c["smtp_porta"] == 587
assert c["smtp_usuario"] == "8f2a10001@smtp-brevo.com"
assert c["smtp_remetente"] == "engenhariadecio2026@gmail.com"
assert mailer.configurado() is True
print("   ✅ Brevo: login do relay e remetente do Gmail convivem")

adm.post("/admin/email", data={
    "acao": "servidor", "smtp_provedor": "sendgrid",
    "smtp_usuario": "qualquer-coisa-que-o-usuario-digitar",
    "smtp_senha": "SG.chave-de-teste",
    "smtp_remetente": "engenhariadecio2026@gmail.com",
    "smtp_ativo": "1"}, follow_redirects=True)
c = email_config.configuracao()
assert c["smtp_usuario"] == "apikey", f"deveria forçar 'apikey': {c['smtp_usuario']}"
assert c["smtp_host"] == "smtp.sendgrid.net"
print("   ✅ SendGrid: usuário forçado para 'apikey', mesmo se digitarem outra coisa")

r = adm.get("/admin/email").data.decode()
for termo in ["Brevo", "SendGrid", "sem 2FA", "Generate a new SMTP key",
              "verificação em duas etapas"]:
    assert termo in r, f"faltou na tela: {termo}"
print("   ✅ a tela traz o passo a passo de cada provedor")

print("\n" + "=" * 58)
print("✅ CONFIGURAÇÃO DE E-MAIL VALIDADA")
print("=" * 58)
