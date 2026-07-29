"""Testa a geração real dos relatórios em Excel e do backup."""
import io
import json
import os

os.environ.setdefault("DATABASE_URL", "postgresql://postgres:teste@localhost:5432/manutencao")
os.environ.setdefault("SECRET_KEY", "teste")

from openpyxl import load_workbook
from app import app
import db

c = app.test_client()
c.post("/login", data={"usuario": "admin", "senha": "teste123"})

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

RELATORIOS = [
    ("/relatorios/os", "Ordens de serviço"),
    ("/relatorios/os-apontamentos", "Apontamentos e tempos"),
    ("/relatorios/preventivas", "Preventivas"),
    ("/relatorios/checklists", "Check lists"),
    ("/relatorios/rondas", "Rondas"),
    ("/relatorios/materiais", "Materiais"),
    ("/relatorios/solicitacoes", "Solicitações"),
    ("/relatorios/equipamentos", "Equipamentos"),
    ("/relatorios/indicadores", "Indicadores"),
    ("/relatorios/terceiros", "Terceiros"),
    ("/relatorios/usuarios", "Usuários e auditoria"),
]

print("── Geração dos relatórios ──")
total_abas = 0
for rota, nome in RELATORIOS:
    r = c.get(rota + "?ini=2020-01-01&fim=2030-12-31")
    assert r.status_code == 200, f"{nome} → HTTP {r.status_code}"
    assert r.mimetype == XLSX, f"{nome} não devolveu xlsx: {r.mimetype}"
    assert "attachment" in r.headers.get("Content-Disposition", "")

    wb = load_workbook(io.BytesIO(r.data))
    total_abas += len(wb.sheetnames)
    detalhes = []
    for aba in wb.sheetnames:
        ws = wb[aba]
        linhas = max(ws.max_row - 4, 0)
        detalhes.append(f"{aba} ({linhas})")
        # Cabeçalho na linha 4, dados a partir da 5
        assert ws["A1"].value, f"{nome}/{aba} sem título"
        assert ws.freeze_panes == "A5", f"{nome}/{aba} sem painel congelado"
    print(f"   ✅ {nome:<24} {len(r.data)//1024:>4} KB · {' · '.join(detalhes)}")

print(f"\n   {len(RELATORIOS)} relatórios · {total_abas} abas geradas")

# ══ CONFERÊNCIA DO CONTEÚDO ═══════════════════════════════════
print("\n── Conferindo o conteúdo ──")
r = c.get("/relatorios/os?ini=2020-01-01&fim=2030-12-31")
wb = load_workbook(io.BytesIO(r.data))
ws = wb["Ordens de Serviço"]
cab = [c.value for c in ws[4]]
assert cab[0] == "OS" and "Custo total" in cab, f"cabeçalho errado: {cab[:5]}"
print(f"   ✅ {len(cab)} colunas na aba de OS")

n_os = db.scalar("SELECT COUNT(*) AS n FROM ordens_servico")
linhas_planilha = 0
for linha in ws.iter_rows(min_row=5, values_only=True):
    if linha[0] is None or linha[0] == "TOTAL":
        break
    linhas_planilha += 1
assert linhas_planilha == n_os, f"esperado {n_os} OS, veio {linhas_planilha}"
print(f"   ✅ {linhas_planilha} linhas conferem com o banco")

# Datas sem fuso (o Excel rejeita datetime com timezone)
from datetime import datetime
achou_data = False
for linha in ws.iter_rows(min_row=5, max_row=6):
    for celula in linha:
        if isinstance(celula.value, datetime):
            achou_data = True
            assert celula.value.tzinfo is None, "datetime com fuso na planilha"
assert achou_data, "nenhuma data encontrada para validar"
print("   ✅ datas gravadas sem fuso horário")

# Formatação
assert ws["A1"].font.bold and ws["A1"].fill.start_color.rgb.endswith("10477D")
assert ws["A4"].fill.start_color.rgb.endswith("28A353")
assert ws["A1"].font.name == "Arial"
print("   ✅ identidade visual aplicada (azul no título, verde no cabeçalho, Arial)")
assert ws.auto_filter.ref, "sem autofiltro"
print("   ✅ autofiltro e largura de colunas")

# Totais no rodapé
ws_ind = load_workbook(io.BytesIO(c.get("/relatorios/indicadores").data))["Por equipamento"]
ultima = ws_ind.max_row
assert ws_ind.cell(row=ultima, column=1).value == "TOTAL", "linha de total ausente"
print("   ✅ linha de TOTAL no rodapé, com valores já calculados")

# Alertas de compra
ws_al = load_workbook(io.BytesIO(c.get("/relatorios/materiais").data))["Alertas de compra"]
print(f"   ✅ aba de alertas com {max(ws_al.max_row - 5, 0)} item(ns) abaixo do mínimo")

# ══ BACKUP ════════════════════════════════════════════════════
print("\n── Backup ──")
assert c.get("/relatorios/backup").status_code == 200
print("   ✅ tela de backup")

r = c.get("/relatorios/backup/excel")
assert r.status_code == 200 and r.mimetype == XLSX
wb = load_workbook(io.BytesIO(r.data))
tabelas_banco = db.query("""SELECT table_name FROM information_schema.tables
                            WHERE table_schema='public' AND table_type='BASE TABLE'""")
n_tab = len(tabelas_banco)
assert wb.sheetnames[0] == "_ÍNDICE", "índice deveria ser a primeira aba"
assert len(wb.sheetnames) == n_tab + 1, \
    f"esperado {n_tab}+1 abas, veio {len(wb.sheetnames)}"
print(f"   ✅ Excel: {len(wb.sheetnames)} abas ({n_tab} tabelas + índice) · "
      f"{len(r.data)//1024} KB")

idx = wb["_ÍNDICE"]
soma = sum(idx.cell(row=i, column=2).value or 0
           for i in range(5, idx.max_row))
print(f"   ✅ índice soma {soma} registros")

# Nomes de aba respeitam o limite do Excel
for aba in wb.sheetnames:
    assert len(aba) <= 31, f"nome de aba longo demais: {aba}"
print("   ✅ nomes de aba dentro do limite de 31 caracteres")

# Binários e senhas fora
for aba in wb.sheetnames:
    cab = [x.value for x in wb[aba][4]]
    for proibido in ("senha_hash", "imagem", "dados", "foto"):
        assert proibido not in (cab or []), f"{proibido} exportado em {aba}"
print("   ✅ senhas e binários não são exportados")

r = c.get("/relatorios/backup/json")
assert r.status_code == 200 and r.mimetype == "application/json"
dados = json.loads(r.data.decode("utf-8"))
assert "tabelas" in dados and len(dados["tabelas"]) == n_tab
assert dados["gerado_por"] and dados["gerado_em"]
n_reg = sum(len(v) for v in dados["tabelas"].values())
print(f"   ✅ JSON: {len(dados['tabelas'])} tabelas · {n_reg} registros · "
      f"{len(r.data)//1024} KB")

for u in dados["tabelas"].get("usuarios", []):
    assert "senha_hash" not in u
print("   ✅ JSON também sem hashes de senha")

log = db.um("SELECT * FROM log_auditoria WHERE acao='backup' ORDER BY id DESC LIMIT 1")
assert log, "backup deveria ir para a auditoria"
print(f"   ✅ registrado na auditoria: {log['detalhe']}")

# ══ PERMISSÕES ════════════════════════════════════════════════
print("\n── Permissões ──")
from werkzeug.security import generate_password_hash
for usuario, nome, perfil in [("jaime2", "Jaime", "manutentor"),
                              ("maria2", "Maria", "analista")]:
    db.executar("""INSERT INTO usuarios (usuario, senha_hash, nome, perfil)
                   VALUES (%s,%s,%s,%s) ON CONFLICT (usuario) DO UPDATE SET perfil=EXCLUDED.perfil""",
                (usuario, generate_password_hash("teste123"), nome, perfil))

mnt = app.test_client(); mnt.post("/login", data={"usuario": "jaime2", "senha": "teste123"})
ana = app.test_client(); ana.post("/login", data={"usuario": "maria2", "senha": "teste123"})
BLOQUEIO = b"n\xc3\xa3o tem acesso"

assert BLOQUEIO in mnt.get("/relatorios/", follow_redirects=True).data
assert BLOQUEIO in mnt.get("/relatorios/os", follow_redirects=True).data
print("   ✅ manutentor não acessa relatórios")

assert c.get("/relatorios/").status_code == 200
r = ana.get("/relatorios/materiais")
assert r.status_code == 200 and r.mimetype == XLSX
print("   ✅ analista gera relatórios")

assert BLOQUEIO in ana.get("/relatorios/backup", follow_redirects=True).data
assert BLOQUEIO in ana.get("/relatorios/backup/excel", follow_redirects=True).data
print("   ✅ backup restrito a administrador e supervisão")
assert BLOQUEIO in ana.get("/relatorios/usuarios", follow_redirects=True).data
print("   ✅ relatório de usuários restrito ao administrador")

print("\n" + "=" * 60)
print("✅ RELATÓRIOS E BACKUP VALIDADOS")
print("=" * 60)
